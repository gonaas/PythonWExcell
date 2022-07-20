from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
	"Joe": {
		"math": 1,
		"science": 1,
		"english": 1,
		"gym": 1
	},
	"Bill": {
		"math": 2,
		"science": 2,
		"english": 2,
		"gym": 2
	},
	"Tim": {
		"math": 3,
		"science": 3,
		"english": 3,
		"gym": 4
	},
	"Sally": {
		"math": 5,
		"science": 5,
		"english": 5,
		"gym": 5
	},
	"Jane": {
		"math": 6,
		"science": 6,
		"english": 6,
		"gym": 6
	}
}

wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ['Name'] + list(data['Joe'].keys())
ws.append(headings)

for person in data:
	grades = list(data[person].values())
	ws.append([person] + grades)

for col in range(2, len(data['Joe'])+2):
	char = get_column_letter(col)
	ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"
 
for col in range(1, 6):
	ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

wb.save("NewGrades.xlsx")