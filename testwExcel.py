from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


#wb=Workbook() #make
wb = Workbook()

ws=wb.active
ws.title = "Data" #name hoja
 
for row in range(1, 10):
     for col in range(1,5):
         char = get_column_letter(col)
         ws[char+str(row)] = char+str(row)
         
wb.save("holita.xlsx")       
    
